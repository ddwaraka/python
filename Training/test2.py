import xlsxwriter
import openpyxl


wb = openpyxl.load_workbook('demo.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

#print sheet.cell(row=2, column=2).value


age_and_names = {}


for i in range(1,11):
    
    age = str(sheet.cell(row=i, column=2).value)
    #print type(age)
    name = sheet.cell(row=i, column=1).value 
    
    if not age in age_and_names:
        age_and_names[age]=[]

    age_and_names[age].append(name)      
       

for key in age_and_names:
    #print key
    print age_and_names[key]
    
    
print age_and_names