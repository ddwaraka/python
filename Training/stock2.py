import requests
import sys
from openpyxl import load_workbook
import json
from collections import defaultdict
import smtplib
import datetime
from threading import Timer

#Reads stock info from Excel and converts into a dict web_url_vs_target_price
def read_excel(file_name):
    web_url_vs_target_price = {}
    wb = load_workbook(filename = file_name, read_only=True)
    ws = wb['Sheet1'] # ws is now an IterableWorksheet
    for row in ws.rows:  
        web_url, target_price = [cell.value for cell in row]
        web_url_vs_target_price[web_url] = target_price
    
    for item in web_url_vs_target_price:
        web_url_vs_target_price[item] = float(web_url_vs_target_price[item])    
    
    return web_url_vs_target_price


#Reads userinfo from Excel and converts it into a dict user_info
def read_user_data():
    user_info = {}
    wb = load_workbook(filename = 'user_details.xlsx', read_only=True)
    ws = wb['Sheet1'] # ws is now an IterableWorksheet
    for row in ws.rows:  
        user_name, password = [cell.value for cell in row]
        user_info[user_name] = password
    return user_info
 
def fetch_price(web_url):
    json_string = requests.get(web_url).content
    json_dict = json.loads(json_string)
    return json_dict['LastTradedPrice'] 

def send_email(email, password):
    subject = "Hello!"
    body = "This message was sent with Python's smtplib."
    message = 'Subject: %s\n\n%s' % (subject, text)
    server = smtplib.SMTP('smtp.gmail.com', 587)
    server.ehlo()
    server.starttls()
    server.login('email', 'password')
    #email and email refer to sender and receiver respectively.
    #Sender and receiver is the same person in this case.
    server.sendmail('email', 'email', message) 
    server.quit()    

def user_auth(user_info):
    if email in user_info:
        if password == user_info[email]:
            print "Success!"
            return True
            
    else:
        print "Wrong Credentials. Please try again"
        sys.exit(-1)

if __name__ == "__main__":
    
    if not len(sys.argv) == 4:
        print 'Invalid number of arguments. Syntax: python stock2.py <input-excel-path> <gmail-username> <gmail-password>'
        sys.exit(-1)
 
    file_name, email, password = sys.argv[1], sys.argv[2], sys.argv[3]
    
    #TODO: Go ahead only if username and password match
    user_info = read_user_data()
    
    if user_auth(user_info) == True:
        web_url_vs_target_price = read_excel(file_name)
        for each_url in web_url_vs_target_price:
            ltp = fetch_price(each_url)
    
            if ltp > web_url_vs_target_price[each_url]:
                #send_email(email, password)
                pass    
    
    
#print read_user_data()




    
    
    
    
        
    
    
    

        